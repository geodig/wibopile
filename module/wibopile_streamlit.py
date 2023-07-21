# -*- coding: utf-8 -*-
"""
Created on Wed Jun 14 12:19:09 2023

@author: YOGB
"""
import openpyxl
import pandas as pd
import numpy as np
import math
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import matplotlib.text as mtext
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from scipy.interpolate import interp1d
import textwrap
from fpdf import FPDF
import os
import folium
import utm

CODE_DIR = r"D:\INTERN\Giovanni Serena Siahaan\streamlit_gio\wibopile_230718\wibopile\module"

class PileCatalog():
    
    def __init__(self):
        self.input_file = os.path.join(CODE_DIR, 'pile_catalog.xlsx')
        
    def get_catalog_table(self):
        """
        Generate a dataframe of pile catalog table.

        Returns
        -------
        df_comb : dataframe
            A dataframe of pile catalog table.

        """
        
        df = pd.read_excel(self.input_file, sheet_name="Sheet1")
        
        return df
    
    def get_catalog_properties(self, catalog_table):
        
        df = catalog_table

        AREA, AREA_SECTION, WEIGHT = [],[],[]
        for i in range(len(df)):
            
            pilediameter = df["Diameter (m)"].iloc[i]
            wallthick = df["Wall thickness (m)"].iloc[i]
            endbearing = df["End bearing"].iloc[i]
            piletype = df["Type"].iloc[i]
            pilematerial = df["Material"].iloc[i]
            
            area_circular = 0.25 * math.pi * (pilediameter)**2
            area_square = (pilediameter)**2
            areafull_circularhollow = 0.25 * math.pi * (pilediameter)**2
            areahol_circularhollow = 0.25 * math.pi * (pilediameter - 2 * wallthick)**2
            area_circularhollow = areafull_circularhollow - areahol_circularhollow
         
            if endbearing == 'with':
                area_circularhollow_used = areafull_circularhollow
            elif endbearing == 'without':
                area_circularhollow_used = area_circularhollow

            if piletype == 'circ_hollow':
                Area = area_circularhollow_used
            elif piletype == 'circular':
                Area = area_circular
            elif piletype == 'square':
                Area = area_square
                
            if piletype == 'circ_hollow':
                Area_sec = area_circularhollow
            elif piletype == 'circular':
                Area_sec = area_circular
            elif piletype == 'square':
                Area_sec = area_square
                
            if pilematerial == 'concrete':
                gamma = 24
            if pilematerial == 'steel':
                gamma = 78.5
            
            if piletype == 'circular':
                weight = (area_circular * gamma)
            elif piletype == 'circ_hollow':
                weight= (area_circularhollow * gamma)
            elif piletype == 'square':
                weight = (area_square * gamma)
            
            AREA.append(Area)
            AREA_SECTION.append(Area_sec)
            WEIGHT.append(weight)
        
        kamus = {
            "Area" : AREA,
            "Area_middle" : AREA_SECTION,
            "Weight" : WEIGHT,
            }
        
        df_prop = pd.DataFrame(kamus)
        
        df_comb = pd.concat([df, df_prop], axis=1)
        
        return df_comb
            

class PileSoilProfile():
    
    def __init__(self, input_file):
        self.input_file = input_file
    
    def read_NSPT(self):
        df = pd.read_excel(self.input_file, sheet_name='nspt')
        
        return df
    
    def read_lithology(self):
        df = pd.read_excel(self.input_file, sheet_name='lithology', usecols=[0,1])
        
        return df
    
    def read_general(self) -> dict:
        wb = openpyxl.load_workbook(self.input_file, data_only=True)
        sh = wb["general"]
        
        col1, col2 = [],[]
        for i in range(sh.max_row):
            col1.append(sh.cell(i+1,1).value)
            col2.append(sh.cell(i+1,2).value)
        
        kamus = {}
        for i in range(len(col1)):
            kamus[col1[i]] = col2[i]
        
        return kamus
    
    def get_general_table(self):
        kamus = self.read_general()
        
        kamus_key = list(kamus.keys())
        kamus_value = list(kamus.values())
        
        kamus2 = {}
        for i in range(len(kamus_key)):
            kamus2[kamus_key[i]] = [kamus_value[i]]   
        
        df = pd.DataFrame(kamus2)
        dft = df.transpose()
        
        return dft
    
    def get_map(self):
        kamus = self.read_general()
        
        X = kamus['X']
        Y = kamus['Y']
        UTM_z = kamus['UTM_zone']
        UTM_h = kamus['UTM_NS']
        label = kamus['BH_label']
        
        if UTM_h == "N":
            hemis = True
        elif UTM_h == "S":
            hemis = False
        latlon = utm.to_latlon(X, Y, UTM_z, northern=hemis)
        lat = latlon[0]
        lon = latlon[1]
        
        location = [lat, lon]
        m = folium.Map(location=location,
                       zoom_start=16,
                        tiles="CartoDB positron",
                       )
        folium.Marker([lat, lon], popup=label, icon=folium.Icon(color='salmon')).add_to(m)
        
        return m
    
    def read_lithokey(self) -> dict:
        inputfile = openpyxl.load_workbook(os.path.join(CODE_DIR,"lithokey.xlsx"))
        sh = inputfile['Sheet1']
        
        col1, col2 = [],[]
        for i in range(sh.max_row - 1):
            col1.append(sh.cell(i+2,1).value)
            col2.append(sh.cell(i+2,2).value)
        
        kamus = {}
        for i in range(len(col1)):
            kamus[col1[i]] = col2[i]
        
        return kamus
    
    def convert_lithology(self):
        lithokey = self.read_lithokey()
        litho_table = self.read_lithology()
        
        for i in range(len(litho_table)):
            litho_table.loc[i, 'lithology_spreadsheet'] = lithokey[litho_table.loc[i, 'lithology']]
        
        return litho_table
    
    def get_soil_profile_table(self):
        """
        Read the input borehole spreadsheet and generate a soil profile table used for bearing
        capacity calculation.

        Returns
        -------
        df : dataframe
            Soil profile table.

        """
        
        df = self.convert_lithology()
        
        df.columns = ['depbot', 'lithology', 'soil']
        
        df.drop('lithology', axis = 1)
        
        deptop = []
        for i in range(len(df) - 1):
            deptop.append(df['depbot'].iloc[i])
        
        deptop.insert(0,0)
        
        df['deptop'] = deptop

        gamma = [17]*len(df)
        nc = ['NC']*len(df)
        
        df['UW'] = gamma
        df['state'] = nc
        
        index = []
        for i in range(len(df)):
            index.append(i+1)
        
        df['Layer'] = index

        df = df.reindex(columns = ['Layer', 'deptop', 'depbot', 'soil', 'UW', 'state'])
        
        df.columns = ['Layer','top','bot','soil','unit weight','state']

        return df
    
    def get_df2(self, soil_profile_table):
        df = self.read_NSPT()
        df_soil = soil_profile_table
        
        DT = np.min(df_soil['top'])
        DB = np.max(df_soil['bot'])
        
        start = np.min(DT) 
        stop = np.max(DB) 
        interval = (stop - start) / 100
       
        top = np.arange(start, stop, interval)
        bottom = []
        for i in range(len(top)):
            if i+1<len(top):
                bottom.append(top[i+1])

        df2 = df
        
        df2.columns = ['depth','NSPT']
        df2['No'] = pd.RangeIndex(start=1, stop=len(df2)+1)
        df2 = df2.set_index('No').reset_index() 
                
        df2['depbot'] = np.nan
        
        for i in range(0, len(df2)):
        
            if i < len(df2) - 1:
                df2.loc[i, 'depbot'] = df2.loc[i, 'depth'] + ((df2.loc[i+1, 'depth'] - df2.loc[i, 'depth']) / 2)
          
            else:
                df2.loc[i, 'depbot'] = stop 
        
        df2['deptop'] = np.roll(df2['depbot'], 1)
        df2.loc[0, 'deptop'] = 0

        df2 = df2.reindex(columns=['No','depth','deptop','depbot','NSPT'])
        
        return df2
    
    def assign_liquefaction(self, reference):
        df = reference
        liquefied = ['no']*len(df)
        df_liquefaction = pd.DataFrame()
        df_liquefaction ["depth top (m)"] = df['top']
        df_liquefaction ["depth bottom (m)"] = df['bot']
        df_liquefaction ["liquefied?"] = liquefied
        
        return df_liquefaction
    
    def assign_borehole(self, soil_profile_table):
        kamus = self.read_general()
        df_ref = soil_profile_table
        df2 = self.get_df2(df_ref)
        
        return kamus, df_ref, df2
    
    def get_soil_profile_chart(self, soil_profile_table):
        """
        Generate an interactive soil profile chart.

        Parameters
        ----------
        soil_profile_table : dataframe
            The dataframe generated using method get_soil_profile_table.

        Returns
        -------
        fig : plotly chart
        
        """
        
        df = soil_profile_table
        bottom_depth = df['bot'].iloc[len(df)-1]
        
        wb = openpyxl.load_workbook(self.input_file, data_only=True)
        wb2 = openpyxl.load_workbook(os.path.join(CODE_DIR,"lithokey.xlsx"))
        
        sheet_info = wb['general']
        sheet_nspt = wb['nspt']
        sheet_litho = wb['lithology']
        sheet_litholist = wb2['Sheet1']
                
        # ID = sheet_info.cell(1,2).value
        BH_label = sheet_info.cell(2,2).value
        Z = sheet_info.cell(5,2).value
        water_level = sheet_info.cell(6,2).value
        elev_unit = sheet_info.cell(7,2).value
        
        spt_depth, spt_value, spt_elev = [],[],[]
        spt_n = sheet_nspt.max_row - 1
        for i in range(spt_n):
            spt_depth.append(sheet_nspt.cell(i+2,1).value)
            spt_value.append(sheet_nspt.cell(i+2,2).value)
            spt_elev.append(Z - spt_depth[i])
        
        litho_depth, litho_value, litho_elev, litho_desc = [],[],[],[]
        litho_n = sheet_litho.max_row - 1
        for i in range(litho_n):
            litho_depth.append(sheet_litho.cell(i+2,1).value)
            litho_value.append(sheet_litho.cell(i+2,2).value)
            litho_desc.append(sheet_litho.cell(i+2,3).value)
        
        litho_depth = litho_depth[:-1]
        litho_depth.append(bottom_depth)
            
        for i in range(litho_n):
            litho_elev.append(Z - litho_depth[i])
        
        litho_desc2 = []
        for i in range(len(litho_desc)):
            if litho_desc[i] == None:
                litho_desc2.append("N/A")
            elif litho_desc[i] != None:
                A = textwrap.wrap(litho_desc[i],30)
                litho_desc2.append('<br>'.join(A))
        
        litho_elev = np.array(litho_elev)
        # litho_elev2 = np.concatenate(([Z],litho_elev))
        
        warnalitho, arsirlitho = {},{}
        for i in range(sheet_litholist.max_row - 1):
            lithoid = sheet_litholist.cell(i+2,1).value
            warna = sheet_litholist.cell(i+2,3).value
            arsir = sheet_litholist.cell(i+2,4).value
            warnalitho[lithoid] = warna
            arsirlitho[lithoid] = arsir
        
        litho_depth2 = [i for i in litho_depth]
        litho_depth2.insert(0,0)
        
        thickness = []
        for i in range(len(litho_depth)):
            thickness.append(litho_depth2[i+1] - litho_depth2[i])
            
        warna, arsir = [],[]
        for lit in litho_value:
            warna.append(warnalitho[lit])
            arsir.append(arsirlitho[lit])
            
        arsir2 = []
        for i in range(len(arsir)):
            if arsir[i] != None:
                arsir2.append(arsir[i])
            else:
                arsir2.append('')
        
        nk = np.empty(shape=(len(litho_value),2,1), dtype='object')
        nk[:,0] = np.array(litho_value).reshape(-1,1)
        nk[:,1] = np.array(litho_desc2).reshape(-1,1)
        
        # PLOTTING
        fig = make_subplots(rows=1, cols=2, column_widths=[0.15, 0.85], horizontal_spacing=0.1)

        # PLOT 1
        fig.add_trace(go.Bar(x=[1]*litho_n, y=thickness, marker_color=warna, marker_pattern_shape=arsir2,
                             customdata=nk,
                             hovertemplate='<b>Lithology:</b> %{customdata[0]}<br><b>Description:</b> %{customdata[1]}',
                             #text=litho_desc, 
                             hoverinfo='text', hoverlabel = dict(namelength = -1),
                             name="borelog"), row=1, col=1)
        fig.update_xaxes(color="white", range=[0.6,1.4], row=1, col=1)
        fig.update_yaxes(range=[np.max(litho_depth),0], row=1, col=1)
        
        
        # PLOT 2
        fig.add_trace(go.Bar(x=spt_value, y=spt_elev, orientation="h",marker_color="navy", name="N-SPT"), row=1, col=2)
        fig.add_trace(go.Scatter(x=[-10,60],y=[water_level,water_level],marker_color="blue", name="water level"), row=1, col=2)
        fig.add_trace(go.Scatter(x=[-10,60],y=[Z,Z],marker_color="black", name="ground surface"), row=1, col=2)
        fig.update_xaxes(color="black", range=[0,50], title="N-SPT value", tickvals=[0,10,20,30,40,50],
                         #showline=True, linewidth=1, linecolor='grey', mirror=True, 
                         row=1, col=2)
        fig.update_yaxes(range=[Z - np.max(litho_depth), Z], 
                         showgrid=True, tick0=0, dtick=5,
                         #showline=True, linewidth=1, linecolor='grey', mirror=True, 
                         row=1, col=2)
        
        # GENERAL SETTING
        fig.update_layout(xaxis_title="NSPT value", 
                          yaxis_title="Depth (m) - Elevation (%s)"%(elev_unit),
                          width=500,height=500,plot_bgcolor="white",
                          title=BH_label,
                          title_x=0.12,
                          font_family="Segoe UI",
                          title_font_family = "Segoe UI semibold",
                          margin=dict(l=40, r=30, t=30, b=40),
                          showlegend=True)
        
        return fig
    
    def get_soil_profile_chart_mpl(self):
        
        wb = openpyxl.load_workbook(self.input_file,data_only=True)
        sheet_info = wb['general']
        sheet_nspt = wb['nspt']
        sheet_litho = wb['lithology']
        sheet_litholist = wb['litholist']
        
        ID = sheet_info.cell(1,2).value
        BH_label = sheet_info.cell(2,2).value
        Z = sheet_info.cell(5,2).value
        water_level = sheet_info.cell(6,2).value
        elev_unit = sheet_info.cell(7,2).value
        
        spt_depth, spt_value, spt_elev = [],[],[]
        spt_n = sheet_nspt.max_row - 1
        for i in range(spt_n):
            spt_depth.append(sheet_nspt.cell(i+2,1).value)
            spt_value.append(sheet_nspt.cell(i+2,2).value)
            spt_elev.append(Z - spt_depth[i])
        
        litho_depth, litho_value, litho_elev = [],[],[]
        litho_n = sheet_litho.max_row - 1
        for i in range(litho_n):
            litho_depth.append(sheet_litho.cell(i+2,1).value)
            litho_value.append(sheet_litho.cell(i+2,2).value)
            litho_elev.append(Z - litho_depth[i])
        
        litho_elev = np.array(litho_elev)
        litho_elev2 = np.concatenate(([Z],litho_elev))
        
        warnalitho, arsirlitho = {},{}
        for i in range(sheet_litholist.max_row - 1):
            lithoid = sheet_litholist.cell(i+2,1).value
            warna = sheet_litholist.cell(i+2,2).value
            arsir = sheet_litholist.cell(i+2,3).value
            warnalitho[lithoid] = warna
            arsirlitho[lithoid] = arsir
        
        some_list,legend_list = [],[]
        litholist = list(warnalitho)
        
        class LegendTitle(object):
            def __init__(self, text_props=None):
                self.text_props = text_props or {}
                super(LegendTitle, self).__init__()
        
            def legend_artist(self, legend, orig_handle, fontsize, handlebox):
                x0, y0 = handlebox.xdescent, handlebox.ydescent
                title = mtext.Text(x0, y0, orig_handle,  **self.text_props)
                handlebox.add_artist(title)
                return title
        
        bottom = np.max(spt_depth)
        
        if bottom <= 50:
            window = (6,6)
        elif bottom > 50 and bottom <= 75:
            window = (6,6)
        elif bottom > 75 and bottom <= 100:
            window = (6,6)
        elif bottom > 100:
            window = (6,6)
        
        # window = (6,6)
        
        # PLOTTING ============================================================
        
        fig,ax = plt.subplots(figsize=window)
        ax.plot([0,100],[Z,Z],color='black',label='ground surface')
        ax.plot([0,100],[water_level,water_level],color='blue',label='water level')
        
        ax.plot([27,27],[Z,np.min(spt_elev)],color='black')        
        ax.barh(spt_elev,spt_value,left=27,color='black',edgecolor='black',zorder=3)
        for j in range(len(spt_value)):
            ax.annotate(str(spt_value[j]), (spt_value[j]+2+27,spt_elev[j]-0.25), color='black',fontsize=11)
            ax.annotate(str(spt_depth[j]), (10,spt_elev[j]-0.2), horizontalalignment = 'right',fontsize=10)
            ax.plot([11.5,15],[spt_elev[j],spt_elev[j]],color='black')
        
        for j in range(len(litho_value)):
            ax.add_patch(patches.Rectangle((15,litho_elev2[j]), 
                                            10, litho_elev2[j+1]-litho_elev2[j], 
                                            # label=litho_value2[i][j] if litho_value2[i][j] not in some_list else '', 
                                            color=warnalitho[litho_value[j]], ec='black', 
                                            hatch=arsirlitho[litho_value[j]]))
    
            if litho_value[j] not in some_list:
                legend_list.append(litho_value[j])
            
            some_list.append(litho_value[j])
        
        legend_index = []    
        for i in range(len(legend_list)):
            legend_index.append(litholist.index(legend_list[i]))
        
        legend_index_sorted = sorted(legend_index)
        
        legend_index2 = []
        for i in range(len(legend_list)):
            legend_index2.append(legend_index_sorted.index(legend_index[i]))
            
        legend_list_sorted = []
        for i in range(len(legend_list)):
            indexs = legend_index2.index(i)
            legend_list_sorted.append(legend_list[indexs])
        
        for i in range(len(legend_list)):
            ax.add_patch(patches.Rectangle((0,-200),1,1,color=warnalitho[legend_list_sorted[i]],
                                            hatch=arsirlitho[legend_list_sorted[i]],ec='black',
                                            label=legend_list_sorted[i]))
        
        # a = nstratitype
        h,l = ax.get_legend_handles_labels()
        
        ax.legend(['Legend'] + h[:2] + ['','Lithology']    + h[2:], 
                  ['']       + l[:2] + ['','']             + l[2:],
                    handler_map={str: LegendTitle({'fontsize': 13})},
                  bbox_to_anchor=(1,1),edgecolor='none',loc="upper left")
        
        # ax.legend(bbox_to_anchor=(1,1), loc="upper left",edgecolor='none')
    
        # ax.text(18,Z+1,BH_label,rotation=90,size=13)
        # ax.text(30,Z+1,'N-SPT value',rotation=0,size=11)
        
        ax.set_title(BH_label+" (N-SPT value)")
        ax.set_xlim(0,85)
        ax.set_ylim(np.min(litho_elev)-1,np.max([Z,water_level])+1)
        ax.set_ylabel('elevation (%s)'%(elev_unit))
        ax.set_xticks([])
        plt.rcParams["axes.grid.axis"] ="y"
        plt.rcParams["axes.grid"] = True
        ax.grid(linewidth=0.3,color='silver')
        plt.tight_layout()
        # plt.show()
        outputname = 'chart_borelog_mpl.png'
        # os.chdir(self.outputdir)
        plt.savefig(outputname, dpi=300)
        
        return fig
    
class PileCalculator():
    
    def __init__(self, borehole):
        self.name = "calculator"
        self.borehole = borehole
        
    
    def calculate_bearing_capacity(self,
                                   pile_parameter_table,
                                   SF_compression,
                                   SF_tension,
                                   cohesive_approach,
                                   granular_approach_shaft,
                                   granular_approach_endB
                                   ):
        """
        Calculate the bearing capacity of given pile parameter and given soil profile.

        Parameters
        ----------
        soil_profile_table : dataframe
            The dataframe generated using method get_soil_profile_table.
        pile_parameter_table : dataframe
            The dataframe generated using class PileCatalog and method get_catalog_table, after filtering.
        SF_compression : float
            Safety factor value for compression loading.
        SF_tension : float
            Safety factor value for tension loading.
        cohesive_approach : string
            Approach to calculate skin friction for cohesive soils, options = "API", "NAVFAC".
        granular_approach_shaft : string
            Approach to calculate skin friction for granular soils, options = "API", "NAVFAC", "OCDI(NSPT)".
        granular_approach_endB : string
            Approach to calculate end bearing for granular soils, options = "API", "NAVFAC".

        Returns
        -------
        df1 : dataframe
            A dataframe containing the calculated ultimate and allowed bearing capacity over depth.

        """
        
        self.pilematerial = pile_parameter_table["Material"].values[0]
        self.piletype = pile_parameter_table["Type"].values[0]
        self.pilediameter = pile_parameter_table["Diameter (m)"].values[0]
        self.wallthick = pile_parameter_table["Wall thickness (m)"].values[0]
        self.endbearing = pile_parameter_table["End bearing"].values[0]
        self.Area = pile_parameter_table["Area"].values[0]
        self.weight = pile_parameter_table["Weight"].values[0]
        
        self.SF_compression = SF_compression
        self.SF_tension = SF_tension
        
        self.cohesiveapproach = cohesive_approach
        self.granularapproachshaft = granular_approach_shaft
        self.granularapproachendB = granular_approach_endB
        
        df1_referensi = self.borehole[1]
        
        DT = np.min(df1_referensi['top'])
        DB = np.max(df1_referensi['bot'])
        
        start = np.min(DT) 
        stop = np.max(DB) 
        interval = (stop - start) / 100
       
        top = np.arange(start, stop+interval, interval)
        
        df1 = pd.DataFrame({'top': top})
        df1['No'] = pd.RangeIndex(start=0, stop=len(df1))
        
        df2 = self.borehole[2]
        
        kamus = self.borehole[0]
        
        GSE = kamus['Z']
        GWD = kamus['water_level']
        
        for i, row in df1.iterrows():
            for j, ref_row in df1_referensi.iterrows():
                if row['top'] >= ref_row['top'] and row['top'] < ref_row['bot']:
                    df1.at[i, 'soil'] = ref_row['soil']

        for i, row in df1.iterrows():
            for j, ref_row in df1_referensi.iterrows():
                if row['top'] >= ref_row['top'] and row['top'] < ref_row['bot']:
                    df1.at[i, 'botsoil'] = ref_row['bot']

        for i, row in df1.iterrows():
            for j, ref_row in df2.iterrows():
                if row['top'] >= ref_row['deptop'] and row['top'] < ref_row['depbot']:
                    df1.at[i, 'nspt'] = ref_row['NSPT']
                elif row['top'] > ref_row['NSPT']:
                    df1.at[i, 'nspt'] = ref_row['NSPT'].max()
        
        for i in range(0, len(df1)):
            if df1.loc[i,'soil'] =='silt' or df1.loc[i,'soil'] == 'clay':
                df1.loc[i,'type'] = 'cohesive'
            else:
                df1.loc[i,'type'] = df1.loc[i,'soil']
        
        for i, row in df1.iterrows():
            for j, ref_row in df2.iterrows():
                if row['top'] >= ref_row['deptop'] and row['top'] < ref_row['depbot']:
                    df1.at[i, 'nspt'] = ref_row['NSPT']
        
        if (df1['botsoil'].iloc[0]+1) > df1['botsoil'].iloc[0]:
                df1['nspt_filter'] =  df1['nspt'] 
                
                
        for i, row in df1.iterrows():
            for j, ref_row in df1_referensi.iterrows():
                if row['top'] >= ref_row['top'] and row['top'] < ref_row['bot']:
                    df1.at[i, 'gamma'] = ref_row['unit weight']
                    
        df1['bot'] = np.roll(df1['top'],-1) 
        if df1['bot'].iloc[-1] == 0:
            df1 = df1.drop(df1.tail(1).index)
        
        df1['elev'] = GSE - df1['top'] #input GSE
        df1.loc[0,'sigtot'] = ((df1.loc[0, 'bot']- df1.loc[0, 'top']) *df1.loc[0, 'gamma'])
        for i in range (1,len(df1)):
            df1.loc[i,'sigtot'] = ((df1.loc[i, 'bot']- df1.loc[i, 'top']) *df1.loc[i, 'gamma']) + df1.loc[i-1, 'sigtot']
            
        df1.loc[0, 'u'] = 0
        
        for i in range(1, len(df1)):
              if df1.loc[i, 'top'] < GWD: 
                  df1.loc[i, 'u'] = 0
              else:
                  df1.loc[i, 'u'] = (df1.loc[i, 'bot'] - df1.loc[i, 'top']) * 10 + df1.loc[i-1, 'u']
        
        df1['sigeff'] = df1['sigtot']-df1['u']
        
        for i in range(0,len(df1)):
            if df1.loc[i,'type'] == 'cohesive':
                df1.loc[i,'su'] = 5* df1.loc[i, 'nspt']
            else:
                df1.loc[i,'su'] = ''
        
        df1 = df1.reindex(columns=['No','top','bot','soil','nspt','type','botsoil','nspt_filter','elev','gamma','sigtot','u','sigeff','su']) 
            

        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None) 
        
        df1['API_su/sigeff'] = np.nan
        df1['API_alpha'] =np.nan
        df1['API_alpha2'] = np.nan
        df1['NAVFAC_alpha_up'] = np.nan
        df1['NAVFAC_alpha_low'] = np.nan
        df1['NAVFAC_su_up'] = np.nan
        df1['NAVFAC_su_low'] = np.nan
        df1['NAVFAC_alpha2'] = np.nan
        df1['selected_alpha'] = np.nan
        df1['cohesive_f'] = np.nan
        
        for i in range(0,len(df1)):
            if df1.loc[i,'type'] == 'cohesive':
                df1.loc[i,'API_su/sigeff'] = round(df1.loc[i,'su']/df1.loc[i,'sigeff'],2)
            else:
                df1.loc[i,'API_su/sigeff'] = ''
        
        for i in range(0,len(df1)):
            if df1.loc[i,'type'] == 'cohesive':
                if df1.loc[i,'API_su/sigeff'] <= 1:
                    df1.loc[i,'API_alpha'] = round(0.5* (df1.loc[i,'API_su/sigeff'])**(-0.5),2)
                else:
                    df1.loc[i,'API_alpha'] = round(0.5* (df1.loc[i,'API_su/sigeff'])**(-0.25),2)
            else:
                df1.loc[i,'API_alpha'] = '' 
        
        for i in range(0,len(df1)):
            if df1.loc[i,'type'] == 'cohesive':
                if df1.loc[i,'API_alpha'] >1:
                    df1.loc[i,'API_alpha2'] = 1
                else:
                    df1.loc[i,'API_alpha2'] = round(df1.loc[i,'API_alpha'],2)
            else: 
                df1.loc[i,'API_alpha2'] = ''
                
        assign_pile_material = self.pilematerial
        for i in range(0, len(df1)):
            if df1.loc[i, 'type'] == 'cohesive':
                if assign_pile_material == 'concrete':
                    if df1.loc[i, 'su'] >= 0 and df1.loc[i, 'su'] < 12:
                        df1.loc[i, 'NAVFAC_alpha_up'] = 1
                    elif df1.loc[i, 'su'] >= 12 and df1.loc[i, 'su'] < 24:
                        df1.loc[i, 'NAVFAC_alpha_up'] = 1
                    elif df1.loc[i, 'su'] >= 24 and df1.loc[i, 'su'] < 48:
                        df1.loc[i, 'NAVFAC_alpha_up'] = 0.96
                    elif df1.loc[i, 'su'] >= 48 and df1.loc[i, 'su'] < 96:
                        df1.loc[i, 'NAVFAC_alpha_up'] = 0.75
                    elif df1.loc[i, 'su'] >= 96:
                        df1.loc[i, 'NAVFAC_alpha_up'] = 0.48
                elif assign_pile_material == 'steel':
                    if df1.loc[i, 'su'] >= 0 and df1.loc[i, 'su'] < 12:
                        df1.loc[i, 'NAVFAC_alpha_up'] = 1
                    elif df1.loc[i, 'su'] >= 12 and df1.loc[i, 'su'] < 24:
                        df1.loc[i, 'NAVFAC_alpha_up'] = 1
                    elif df1.loc[i, 'su'] >= 24 and df1.loc[i, 'su'] < 48:
                        df1.loc[i, 'NAVFAC_alpha_up'] = 0.92
                    elif df1.loc[i, 'su'] >= 48 and df1.loc[i, 'su'] < 96:
                        df1.loc[i, 'NAVFAC_alpha_up'] = 0.70
                    elif df1.loc[i, 'su'] >= 96:
                        df1.loc[i, 'NAVFAC_alpha_up'] = 0.36  
            elif df1.loc[i, 'type'] != 'cohesive':
                df1.loc[i, 'NAVFAC_alpha_up'] = ''
                
        for i in range(0, len(df1)):
            if df1.loc[i, 'type'] == 'cohesive':
                if assign_pile_material == 'concrete':
                    if df1.loc[i, 'su'] >= 0 and df1.loc[i, 'su'] < 12:
                        df1.loc[i, 'NAVFAC_alpha_low'] = 1
                    elif df1.loc[i, 'su'] >= 12 and df1.loc[i, 'su'] < 24:
                        df1.loc[i, 'NAVFAC_alpha_low'] = 0.96
                    elif df1.loc[i, 'su'] >= 24 and df1.loc[i, 'su'] < 48:
                        df1.loc[i, 'NAVFAC_alpha_low'] = 0.75
                    elif df1.loc[i, 'su'] >= 48 and df1.loc[i, 'su'] < 96:
                        df1.loc[i, 'NAVFAC_alpha_low'] = 0.48
                    elif df1.loc[i, 'su'] >= 96:
                        df1.loc[i, 'NAVFAC_alpha_low'] = 0.33
                if assign_pile_material == 'steel':
                  if df1.loc[i, 'su'] >= 0 and df1.loc[i, 'su'] < 12:
                      df1.loc[i, 'NAVFAC_alpha_low'] = 1
                  elif df1.loc[i, 'su'] >= 12 and df1.loc[i, 'su'] < 24:
                      df1.loc[i, 'NAVFAC_alpha_low'] = 0.92
                  elif df1.loc[i, 'su'] >= 24 and df1.loc[i, 'su'] < 48:
                      df1.loc[i, 'NAVFAC_alpha_low'] = 0.70
                  elif df1.loc[i, 'su'] >= 48 and df1.loc[i, 'su'] < 96:
                      df1.loc[i, 'NAVFAC_alpha_low'] = 0.36
                  elif df1.loc[i, 'su'] >= 96:
                      df1.loc[i, 'NAVFAC_alpha_low'] = 0.19  
          
            elif df1.loc[i, 'type'] != 'cohesive':
                df1.loc[i, 'NAVFAC_alpha_low'] = ''
        
        for i in range(0, len(df1)):
            if df1.loc[i, 'type'] == 'cohesive':
                if df1.loc[i, 'su'] >= 0 and df1.loc[i, 'su'] < 12:
                    df1.loc[i, 'NAVFAC_su_up'] = 12
                elif df1.loc[i, 'su'] >= 12 and df1.loc[i, 'su'] < 24:
                    df1.loc[i, 'NAVFAC_su_up'] = 24
                elif df1.loc[i, 'su'] >= 24 and df1.loc[i, 'su'] < 48:
                    df1.loc[i, 'NAVFAC_su_up'] = 48
                elif df1.loc[i, 'su'] >= 48 and df1.loc[i, 'su'] < 96:
                    df1.loc[i, 'NAVFAC_su_up'] = 96
                elif df1.loc[i, 'su'] >= 96:
                    df1.loc[i, 'NAVFAC_su_up'] = 192       
            else:
                df1.loc[i, 'NAVFAC_su_up'] = ''
                
        
        for i in range(0, len(df1)):
            if df1.loc[i, 'type'] == 'cohesive':
                if df1.loc[i, 'su'] >= 0 and df1.loc[i, 'su'] < 12:
                    df1.loc[i, 'NAVFAC_su_low'] = 0
                elif df1.loc[i, 'su'] >= 12 and df1.loc[i, 'su'] < 24:
                    df1.loc[i, 'NAVFAC_su_low'] = 12
                elif df1.loc[i, 'su'] >= 24 and df1.loc[i, 'su'] < 48:
                    df1.loc[i, 'NAVFAC_su_low'] = 24
                elif df1.loc[i, 'su'] >= 48 and df1.loc[i, 'su'] < 96:
                    df1.loc[i, 'NAVFAC_su_low'] = 48
                elif df1.loc[i, 'su'] >= 96:
                    df1.loc[i, 'NAVFAC_su_low'] = 96       
            else:
                df1.loc[i, 'NAVFAC_su_low'] = ''    

        #alpha2-----------------------------------------------------------------------
        for i in range (0, len (df1)):
            a = df1.loc[i,'su']
            b = df1.loc[i,'NAVFAC_su_low']
            c= df1.loc[i,'NAVFAC_su_up']
            d= df1.loc [i, 'NAVFAC_alpha_up']
            e = df1.loc[i,'NAVFAC_alpha_low']
            if df1.loc [i,'type'] == 'cohesive':
                df1.loc[i, 'NAVFAC_alpha2'] = round(((a-b)/(c-b))*(d-e)+e,2)
            else: 
                df1.loc[i, 'NAVFAC_alpha2'] = ''
                
        
        cohesive_approach_assign = self.cohesiveapproach
        for i in range(0, len(df1)):
            if df1.loc[i, 'type'] == 'cohesive':
                if cohesive_approach_assign == 'API': 
                    df1['selected_alpha'].iloc[i] = df1['API_alpha2'].iloc[i]
                elif cohesive_approach_assign == 'NAVFAC':
                    df1['selected_alpha'].iloc[i] = df1['NAVFAC_alpha2'].iloc[i]
        
        for i in range(0, len(df1)):
            x = df1.loc[i,'su']
            y = df1.loc[i,'selected_alpha']
            if df1.loc[i, 'type'] == 'cohesive':
                df1.loc[i,'cohesive_f'] = round(x*y,2)
            else:
                df1.loc[i,'cohesive_f'] =''
                    
        
        for i in range(0, len(df1)):
            if df1.loc[i, 'type'] == 'cohesive':
                df1.loc[i,'NAVFAC_phi'] = ''
            else:
                df1.loc[i,'NAVFAC_phi']= round(27.1+0.3*df1.loc[i,'nspt_filter']-0.00054*df1.loc[i,'nspt_filter']**2,2)
        
        typelist = ['sand loose', 'sand medium', 'sand dense', 'sand very dense']
        
        for i in range(0, len(df1)):
            if df1.loc[i, 'type'] not in typelist:
                df1.loc[i,'NAVFAC_lamda'] = ''
            elif df1.loc[i, 'type'] in typelist:
                if assign_pile_material == 'concrete':
                    df1.loc[i,'NAVFAC_lamda'] = round(0.75* df1.loc[i,'NAVFAC_phi'],2)
                elif assign_pile_material == 'steel':
                    df1.loc[i,'NAVFAC_lamda'] = 20
                
        for i in range(0, len(df1)):
            if df1.loc[i, 'type'] == 'cohesive':
                df1.loc[i, 'NAVFAC_beta'] = ''
            else:
                df1.loc[i,'NAVFAC_lamda'] = round(math.radians(df1.loc[i,'NAVFAC_lamda']),2)
                K_driven = 1
                df1.loc[i, 'NAVFAC_beta'] = round(K_driven * math.tan(df1.loc[i, 'NAVFAC_lamda']),2)
                
        for i in range (0, len(df1)):
            if df1.loc[i,'soil'] == 'sand loose':
                df1.loc[i,'API_beta'] = 0.29
            elif df1.loc[i,'soil'] == 'sand medium':
                df1.loc[i,'API_beta'] = 0.37
            elif df1.loc[i,'soil'] == 'sand dense':
                df1.loc[i,'API_beta'] = 0.46
            elif df1.loc[i,'soil'] == 'sand very dense':
                df1.loc[i,'API_beta'] = 0.56
            else:
                df1.loc[i,'API_beta'] = ''
                
        for i in range(0, len(df1)):
            if df1.loc[i, 'type'] == 'cohesive':
                df1.loc[i,'selected_beta'] =''
            else:
                if cohesive_approach_assign == 'API':
                    df1.loc[i,'selected_beta'] = df1.loc[i,'API_beta']
                elif cohesive_approach_assign == 'NAVFAC':
                    df1.loc[i,'selected_beta'] = df1.loc[i,'NAVFAC_beta']
                
        for i in range(0, len(df1)):
            if df1.loc[i, 'type'] == 'cohesive':
                df1.loc[i,'OCDI_f'] = ''
            else:
                df1.loc[i,'OCDI_f'] = 2* df1.loc[i,'nspt']
                
        granular_approach_shaft_assign = self.granularapproachshaft 
        df1.replace({'selected_beta': {'': np.nan}, 'sigeff': {'': np.nan}}, inplace=True) 
        df1.fillna(value={'selected_beta': 0.0, 'sigeff': 0.0}, inplace=True) 
        for i in range(0, len(df1)):
            if df1.loc[i, 'type'] == 'cohesive':
                df1.loc[i,'selected_f'] = 0.0
            else:
                if granular_approach_shaft_assign == 'OCDI':
                    df1.loc[i,'selected_f'] = df1.loc[i,'OCDI_f']
                else:
                    x = float(df1.loc[i,'selected_beta'])
                    y = float(df1.loc[i,'sigeff'])
                    df1.loc[i,'selected_f'] = round(x*y,2)
                
        for i in range(0,len(df1)):
            if df1.loc[i,'soil'] =='silt' or df1.loc[i,'soil'] == 'clay':
                df1.loc[i,'limit_type'] = 'NONE'
            else:
                df1.loc[i,'limit_type'] = df1.loc[i,'soil']
                
        for i in range (0,len(df1)):
            if df1.loc[i,'limit_type'] == 'sand loose':
                df1.loc[i,'f_max'] = 67
                df1.loc[i,'q_max'] = 3000
            elif df1.loc[i,'limit_type'] == 'sand medium':
                df1.loc[i,'f_max'] = 81
                df1.loc[i,'q_max'] = 5000
            elif df1.loc[i,'limit_type'] == 'sand dense':
                df1.loc[i,'f_max'] = 96
                df1.loc[i,'q_max'] = 10000
            elif df1.loc[i,'limit_type'] == 'sand very dense':
                df1.loc[i,'f_max'] = 115
                df1.loc[i,'q_max'] = 12000
            elif df1.loc[i,'limit_type'] == 'NONE':
                df1.loc[i,'f_max'] = 0
                df1.loc[i,'q_max'] = 0
                
        for i in range(0, len(df1)):       
            if df1.loc[i, 'type'] == 'cohesive':
                df1.loc[i,'limittype_f'] = round(df1.loc[i,'cohesive_f'],2)
            else:
                df1.loc[i,'limittype_f'] = round(df1.loc[i,'selected_f'],2)

        for i in range (0, len(df1)):
            if df1.loc[i,'limit_type'] != "NONE":
                if df1.loc[i,'limittype_f'] < df1.loc[i,'f_max']:
                    df1.loc[i,'limittype_fmax'] =  df1.loc[i,'limittype_f']
                else:
                    df1.loc[i,'limittype_fmax'] =  df1.loc[i,'f_max']
            else:
                df1.loc[i,'limittype_fmax'] =  df1.loc[i,'limittype_f']
                
        assign_pile_type = self.piletype
        diameter = self.pilediameter
        
        for i in range (0, len(df1)):
            if assign_pile_type == 'circ_hollow':
                df1.loc[i,'circum'] = 22/7*diameter
            elif assign_pile_type == 'circular':
                df1.loc[i,'circum'] = 22/7*diameter
            elif assign_pile_type == 'square':
                df1.loc[i,'circum'] = diameter**2
            
        for i in range (0, len(df1)):
            df1.loc[i,'compression_qf'] = round(df1.loc[i,'limittype_fmax']*df1.loc[i,'circum']*((df1.loc[i,'bot']-df1.loc[i,'top'])),2)
            
        df1.loc[0,'compression_qf_cum'] = df1.loc[0,'compression_qf']
        for i in range(1, len(df1)):
            df1.loc[i,'compression_qf_cum'] = round(df1.loc[i-1,'compression_qf_cum'] + df1.loc[i,'compression_qf'],2)
              
        for i in range (0, len(df1)):
            if df1.loc[i,'soil'] == 'sand loose':
                df1.loc[i,'nq_API'] = 12
            elif df1.loc[i,'soil'] == 'sand medium':
                df1.loc[i,'nq_API'] = 20
            elif df1.loc[i,'soil'] == 'sand dense':
                df1.loc[i,'nq_API'] = 40
            elif df1.loc[i,'soil'] == 'sand very dense':
                df1.loc[i,'nq_API'] = 50
            else:
                df1.loc[i,'nq_API'] = 9
                
        for i in range(0, len(df1)):       
            if df1.loc[i, 'type'] == 'cohesive':
                df1.loc[i,'nq_NAVFAC'] = 9
            else:
                df1.loc[i,'nq_NAVFAC'] = 0.0739 * math.exp(0.1876*df1.loc[i,'NAVFAC_phi'])
                
        granular_approachendB = self.granularapproachendB 
        for i in range(0, len(df1)): 
            if granular_approachendB == 'API':
                df1.loc[i,'nq_select'] = df1.loc[i,'nq_API']
            else:
                df1.loc[i,'nq_select'] = df1.loc[i,'nq_NAVFAC']
        for i in range(0, len(df1)): 
            if granular_approachendB == 'API':
                df1.loc[i,'nq_select'] = df1.loc[i,'nq_API']
            else:
                df1.loc[i,'nq_select'] = df1.loc[i,'nq_NAVFAC']
                
        for i in range(0, len(df1)):       
            if df1.loc[i, 'type'] == 'cohesive':
                df1.loc[i,'compression_q'] = df1.loc[i,'nq_select']* df1.loc[i,'su']
            else:
                df1.loc[i,'compression_q'] = df1.loc[i,'nq_select']* df1.loc[i,'sigeff']
        
        for i in range(0, len(df1)): 
            area = self.Area
        for i in range (0, len(df1)):
            df1.loc[i,'compression_qp'] = df1.loc[i,'compression_q']*area
            
        
        for i in range (0,len(df1)):
            df1.loc[i,'compression_qult'] = df1.loc[i,'compression_qp']+df1.loc[i,'compression_qf_cum']
            
        SF_comp_assign = SF_compression
        SF_tens_assign = SF_tension
        for i in range (0,len(df1)):
            df1.loc[i,'compression_qall']= df1.loc[i,'compression_qult']/SF_comp_assign
            
        Weight = self.weight
        for i in range (0,len(df1)):
            df1.loc[i,'qult_tens'] = df1.loc[i,'compression_qf_cum']*0.7+Weight*df1.loc[i,'bot']
        
        for i in range (0,len(df1)):
            df1.loc[i,'qall_tens']= df1.loc[i,'qult_tens']/SF_tens_assign
        
        return df1
    
    
    def calculate_bearing_capacity_liquefied(self, reference, calculation, pile_parameter_table, SF_compression, SF_tension):
        self.weight = pile_parameter_table["Weight"].values[0]
        Weight = self.weight
        
        ref = reference
        df = calculation
        
        df1 = pd.DataFrame()
        df1['liquefied?'] = []
        df1['top'] = []
        df1['bot'] = []
        
        i=0
        for k in range(len(df)):
            df1.loc[i,'top'] = df.loc[k,'top']
            df1.loc[i,'bot'] = df.loc[k,'bot']
            
            for j, ref_row in ref.iterrows():
                if (df1.loc[i,'top'] >= ref_row['depth top (m)']) and (df1.loc[i,'top'] < ref_row['depth bottom (m)']):
                    df1.at[i, 'liquefied?'] = ref_row['liquefied?']
                    
                    if df1.loc[i,'liquefied?'] == 'no':                    
                        df1.loc[i,'compression_qf'] = df.loc[k,'compression_qf']
                    else: 
                        df1.loc[i,'compression_qf'] = 0
                        
                    df1.loc[0,'compression_qf_cum'] = df1.loc[0,'compression_qf']

                    for idx in range (1,len(df1)):
                        df1.loc[idx,'compression_qf_cum'] = df1.loc[idx-1,'compression_qf_cum'] + df1.loc[idx,'compression_qf']
                        
                    
                    if df1.loc[i,'liquefied?'] == 'no':
                        df1.loc[i,'compression_qp'] = df.loc[k,'compression_qp']
                    else: 
                        df1.loc[i,'compression_qp'] = 0
                    
                    
                    if (df1.loc[i,'bot'] > ref_row['depth bottom (m)']):
                        df1.loc[i,'bot'] = ref_row['depth bottom (m)']
                        
                        i+=1
                        df1.loc[i,'top'] = ref_row['depth bottom (m)']
                        df1.loc[i,'bot'] = df.loc[k,'bot']
                        
                        if df1.loc[i,'liquefied?'] == 'no':                    
                            df1.loc[i,'compression_qf'] = df.loc[k,'compression_qf']
                        else: 
                            df1.loc[i,'compression_qf'] = 0
                            
                        df1.loc[i,'compression_qf_cum'] = df.loc[k,'compression_qf_cum']
                            
                        
                        if df1.loc[i,'liquefied?'] == 'no':
                            df1.loc[i,'compression_qp'] = df.loc[k,'compression_qp']
                        else: 
                            df1.loc[i,'compression_qp'] = 0
                            
                    for i, row in df1.iterrows(): 
                        
                        df1.loc[i,'compression_qult'] = df1.loc[i,'compression_qf_cum'] + df1.loc[i,'compression_qp']
                        df1.loc[i,'compression_qall'] = df1.loc[i,'compression_qult'] / (SF_compression/1.3)
                        df1.loc[i,'qult_tens'] = (df1.loc[i,'compression_qf_cum']*0.7) + (Weight*df1.loc[i,'bot'])
                        df1.loc[i,'qall_tens'] = df1.loc[i,'qult_tens']/(SF_tension/1.3)
                        
            i+=1     
              
        return df1
            
      
    def get_bearing_capacity_from_depth(self, data_table, depth):
        """
        Generate a table of summary for ultimate and allowable bearing capacity at a certain pile depth.

        Parameters
        ----------
        data_table : dataframe
            Generated from method calculate_bearing_capacity.
        depth : float
            Pile depth of interest.

        Returns
        -------
        result_summary : dataframe
            Result summary table.

        """
        # self.depth = depth
        df = data_table
        list_depth = df["bot"].tolist()
        list_qfcum = df["compression_qf_cum"].tolist()
        list_qp = df["compression_qp"].tolist()
        list_qult_com = df["compression_qult"].tolist()
        list_qall_com = df["compression_qall"].tolist()
        list_qult_ten = df["qult_tens"].tolist()
        list_qall_ten = df["qall_tens"].tolist()
        
        qfcum = interp1d(list_depth, list_qfcum)(depth)
        qp = interp1d(list_depth, list_qp)(depth)
        qult_com = interp1d(list_depth, list_qult_com)(depth)
        qall_com = interp1d(list_depth, list_qall_com)(depth)
        qult_ten = interp1d(list_depth, list_qult_ten)(depth)
        qall_ten = interp1d(list_depth, list_qall_ten)(depth)
        ratio_qf = qfcum / qult_com
        ratio_qp = qp / qult_com
        
        kamus = {
            "Parameter" : ["compression_ultimate", "compression_allowed", "tension_ultimate", "tension_allowed"],
            "Notation" : ["Qult_com", "Qall_com", "Qult_ten", "Qall_ten"],
            "Unit" : ["kN"]*4,
            "Total" : [qult_com, qall_com, qult_ten, qall_ten],
            "Shaft" : [qfcum, ratio_qf * qall_com, None, None],
            "End-bearing" : [qp, ratio_qp * qall_com, None, None],
            }
        
        result_summary = pd.DataFrame(kamus)
        
        return result_summary
        
    def calculate_kh(self, data_table, pile_diameter):
        """
        Calculate the horizontal spring coefficient for all soil layers.

        Parameters
        ----------
        data_table : dataframe
            Generated from method calculate_bearing_capacity.

        Returns
        -------
        df3 : dataframe
            Extension of data_table dataframe, including new columns for horizontal spring calculation.
        df4 : dataframe
            Summary of horizontal spring coefficient for all soil layers.

        """
        
        df = self.borehole[1]
        df3 = pd.DataFrame()
        df1 = data_table
        df3['No'] = df1['No']
        df3['top'] = df1['top']
        df3['bot'] = df1['bot']
        df3['soil'] = df1['soil']
        df3['nspt'] = df1['nspt']
        df3['type'] = np.nan
        df3['Layer']= np.nan
        df3['state'] = np.nan
        
        for i in range (0, len(df3)):
            if df3.loc[i,'soil'] == 'clay':
                df3.loc[i,'type'] = 'clay'        
            elif df3.loc[i,'soil'] == 'silt':
                df3.loc[i,'type'] = 'silt'         
            else:
                df3.loc[i,'type']= 'sand'
        
        df3['botsoil'] = df1['botsoil']
        df3['nspt_filter'] = df1['nspt_filter']
        df3['elev'] = df1['elev']
        df3['gamma'] = df1['gamma']
        df3['sigtot'] = df1['sigtot']
        df3['u'] = df1['u']
        df3['sigeff'] = df1['sigeff']
       
        for i, row in df3.iterrows():
            for j, ref_row in df.iterrows():
                if row['top'] >= ref_row['top'] and row['top'] < ref_row['bot']:
                    df3.at[i, 'Layer'] = ref_row['Layer']
                    
        for i, row in df3.iterrows():
            for j, ref_row in df.iterrows():
                if row['top'] >= ref_row['top'] and row['top'] < ref_row['bot']:
                    df3.at[i, 'state'] = ref_row['state']
        
        #alpha
        for i in range (0, len(df3)):
            if df3.loc[i,'type'] == 'clay' and df3.loc[i,'state'] == 'OC':
                df3.loc[i,'alpha']  = 1
            elif df3.loc[i,'type'] == 'clay' and df3.loc[i,'state'] == 'NC':
                df3.loc[i,'alpha']  = 0.67
            elif df3.loc[i,'type'] == 'clay' and df3.loc[i,'state'] == 'UC':
                df3.loc[i,'alpha']  = 0.67
            elif df3.loc[i,'type'] == 'silt' and df3.loc[i,'state'] == 'OC':
                df3.loc[i,'alpha']  = 0.67
            elif df3.loc[i,'type'] == 'silt' and df3.loc[i,'state'] == 'NC':
                df3.loc[i,'alpha']  = 0.5
            elif df3.loc[i,'type'] == 'silt' and df3.loc[i,'state'] == 'UC':
                df3.loc[i,'alpha']  = 0.5
            elif df3.loc[i,'type'] == 'sand' and df3.loc[i,'state'] == 'OC':
                df3.loc[i,'alpha']  = 0.5
            elif df3.loc[i,'type'] == 'sand' and df3.loc[i,'state'] == 'NC':
                df3.loc[i,'alpha']  = 0.33
            elif df3.loc[i,'type'] == 'sand' and df3.loc[i,'state'] == 'UC':
                df3.loc[i,'alpha']  = 0.33

        #beta hi
        for i in range (0, len(df3)):
            if df3.loc[i,'type'] == 'clay' :
                df3.loc[i,'beta_hi']  = 1
            elif df3.loc[i,'type'] == 'silt' :
                df3.loc[i,'beta_hi']  = 2.5
            elif df3.loc[i,'type'] == 'sand' :
                df3.loc[i,'beta_hi']  = 1.5
            
        #beta lo
        for i in range (0, len(df3)):
            if df3.loc[i,'type'] == 'clay' :
                df3.loc[i,'beta_lo']  = 2.5
            elif df3.loc[i,'type'] == 'silt' :
                df3.loc[i,'beta_lo']  = 3
            elif df3.loc[i,'type'] == 'sand' :
                df3.loc[i,'beta_lo']  = 3

        #Em-hi, Em_lo & Em_avg
        for i in range (0, len(df3)):
            df3.loc[i,'em_hi'] = math.ceil(df3.loc[i,'nspt_filter']/df3.loc[i,'beta_hi']*1000)
            df3.loc[i,'em_lo'] = math.ceil(df3.loc[i,'nspt_filter']/df3.loc[i,'beta_lo']*1000)
            df3.loc[i,'em_avg'] = math.ceil((df3.loc[i,'em_hi'] + df3.loc [i,'em_lo'])/2)
        
        # diameter = self.pilediameter
        diameter = pile_diameter
        R = diameter/2
        R0 = 0.3
        for i in range (0, len(df3)):
                if R < R0:
                    df3.loc[i,'kh_hi'] = 1/((2*R/df3.loc[i,'em_hi'])*(4*2.65**df3.loc[i,'alpha']+3*df3.loc[i,'alpha'])/18)
                else:
                    df3.loc[i,'kh_hi'] = 1/((R/3/df3.loc[i,'em_hi'])*(1.3*R0*(2.65*R/R0)**df3.loc[i,'alpha']+df3.loc[i,'alpha']*R))

        for i in range (0, len(df3)):
            if R < R0:
                df3.loc[i,'kh_lo'] = math.ceil(1/((2*R/df3.loc[i,'em_lo'])*(4*2.65**df3.loc[i,'alpha']+3*df3.loc[i,'alpha'])/18))
            else:
                df3.loc[i,'kh_lo'] = math.ceil(1/((R/3/df3.loc[i,'em_lo'])*(1.3*R0*(2.65*R/R0)**df3.loc[i,'alpha']+df3.loc[i,'alpha']*R)))

        for i in range (0, len(df3)):                              
            if R < R0:
                df3.loc[i,'kh_avg'] = math.ceil(1/((2*R/df3.loc[i,'em_avg'])*(4*2.65**df3.loc[i,'alpha']+3*df3.loc[i,'alpha'])/18))
            else:
                df3.loc[i,'kh_avg'] = math.ceil(1/((R/3/df3.loc[i,'em_avg'])*(1.3*R0*(2.65*R/R0)**df3.loc[i,'alpha']+df3.loc[i,'alpha']*R)))
                     
        df4 = df3.groupby('Layer')[['kh_hi', 'kh_avg', 'kh_lo']].mean().reset_index()
        
        df4.columns = ['Layer', 'kh hi', 'kh med', 'kh lo']
        df4['kh hi [kN/m3]'] = df4['kh hi'].round(1)
        df4['kh med [kN/m3]'] = df4['kh med'].round(1)
        df4['kh lo [kN/m3]'] = df4['kh lo'].round(1)
        # df4.loc [i,'Elevation top [mEGM]'] = df['GSE']-df.loc[i,'bot'] 
        # df4 ['Elevation bot [mEGM]'] = df['GSE']-df.loc[i,'top']
        df4 ['soil type'] = df['soil']        
        df4['No'] = pd.RangeIndex(start=1, stop=len(df4) + 1)
        
        df4 = df4.reindex(columns=['No','Layer','kh hi [kN/m3]','kh med [kN/m3]','kh lo [kN/m3]'])
                
        for i in range (0,len(df4)):
            
            df4.loc[i,'kh pile high [kN/m2]'] = (df4.loc[i,'kh hi [kN/m3]'] * diameter).round(0)                       
            # df4.loc[i,'kh pile high'].round(0)
            df4.loc[i,'kh pile * highv2 [kN/m2]'] = (df4.loc[i,'kh hi [kN/m3]'] * diameter * math.sqrt(2)).round(0)  
            # df4.loc[i,'kh pile * high√2 '].round(0) 
            
            
            df4.loc[i,'kh pile low [kN/m2]'] = (df4.loc[i,'kh lo [kN/m3]'] * diameter).round(0)  
            # df4.loc[i,'kh pile low'].round(0)
            df4.loc[i,'kh pile low/v2 [kN/m2]'] = (df4.loc[i,'kh lo [kN/m3]'] * diameter / math.sqrt(2)).round(0)      
           
            df4.loc[i,'kh pile avg [kN/m2]'] = (df4.loc[i,'kh med [kN/m3]'] * diameter).round(0)  

             
        df4 = df4.reindex(columns=['No','Layer','kh lo [kN/m3]','kh med [kN/m3]','kh hi [kN/m3]','kh pile low/v2 [kN/m2]','kh pile low [kN/m2]','kh pile avg [kN/m2]','kh pile high [kN/m2]','kh pile * highv2 [kN/m2]'])
        
        return df3, df4
    
    def calculate_kv(self, data_table, pile_depth, load_compression, load_tension, area_section, fc_pile = 45):
        """
        Calculate the vertical spring coefficient at the pile bottom tip location.

        Parameters
        ----------
        data_table : dataframe
            Generated using the method calculate_bearing_capacity.
        pile_depth : float
            Pile depth of interest.
        load_compression : float
            Compression load in kN.
        area_section : float
            Area section of the pile.
        fc_pile : int, optional
            Concrete uniaxial compression strength (fc') in MPa. The default is 45.

        Returns
        -------
        df5 : dataframe
            Table of vertical spring coefficient for varying compression load and settlement value.
        df6 : dataframe
            Table of summary for vertical spring coefficient.

        """
        
        df = self.borehole[1]
        df2 = data_table
        self.pile_depth = pile_depth
        self.load_compression = load_compression
        self.load_tension = load_tension
        
        if pile_depth < df['top'].max():
            top_bearing_layer = df.loc[df['top'] < pile_depth, 'top'].max()
        else:
            top_bearing_layer = df['top'].max()
            
                    
        assign_pile_type = self.piletype
        diameter = self.pilediameter
                
        if assign_pile_type == 'square':
            deq = 1.3 * diameter * 1000
        else: 
            deq = diameter * 1000

        L = pile_depth        
        delta_l = pile_depth - top_bearing_layer
        l = L - delta_l

        #interpolasi data
        
        interpolated_qf_cum = interp1d(df2['top'],df2['compression_qf_cum'], kind='linear')
        
        interpolated_qp = interp1d(df2['top'],df2['compression_qp'], kind='linear')
        

        condition = df2['top'] == pile_depth #kalau ga ada dilakukan interpolasi pada hasil else
        if condition.any():
            bearing_capacity_ULS_shaft = df2.loc[condition,'compression_qf_cum'].values[0]
            bearing_capacity_ULS_tip = df2.loc[condition, 'compression_qp'].values[0]
        else:
            bearing_capacity_ULS_shaft = interpolated_qf_cum(pile_depth)
            bearing_capacity_ULS_tip = interpolated_qp(pile_depth)
            
        assign_pile_material = self.pilematerial
        
        if assign_pile_material == 'steel':
            Epile = 20000000  # kPa
        else: 
            Epile = 4700 * (fc_pile)**0.5 * 1000  # kPa
        
        # for i in range (0,len(df)):
        Apile = area_section

        df5 = pd.DataFrame()
        
        sb_data = []
                
        for i in range(0, 41):
            sb_data.append("{:.2f}".format(i/20))
            
        for i in range(21, 31):
            sb_data.append("{:.2f}".format(i/10))
            
        for i in range(4, 26):
            sb_data.append("{:.2f}".format(i)) 
            
        for i in range(30, 110, 10):
            sb_data.append("{:.2f}".format(i))

        df5['sb'] = sb_data
        df5['sb'] = df5['sb'].astype(float)

        for i in range(len(df5)):
            df5.loc[i, 'sb/deq'] = (df5.loc[i,'sb']) / deq *100
            
        for i in range(len(df5)):
            if df5.loc[i, 'sb'] <= 0.47:
                df5.loc[i, 'Fs/Fmax;s'] = 52 * df5.loc[i, 'sb']
                df5.loc[i, 'Fp/Fmax;p'] = 52 * df5.loc[i, 'sb/deq']
            elif df5.loc[i, 'sb'] < 10.6:
                df5.loc[i, 'Fs/Fmax;s'] = 42.9 + (24.5 * math.log(df5.loc[i, 'sb']))
                df5.loc[i, 'Fp/Fmax;p'] = 42.9 + (24.5 * math.log(df5.loc[i, 'sb/deq']))
            else:
                df5.loc[i, 'Fs/Fmax;s'] = 100
                df5.loc[i, 'Fp/Fmax;p'] = 100
                
        for i in range(len(df5)):
            if df5.loc[i, 'sb/deq'] <= 0.47:
                df5.loc[i, 'Fp/Fmax;p'] = 52 * df5.loc[i, 'sb/deq']
            elif df5.loc[i, 'sb/deq'] < 10.6:
                df5.loc[i, 'Fp/Fmax;p'] = 42.9 + (24.5 * math.log(df5.loc[i, 'sb/deq']))
            else:   
                df5.loc[i, 'Fp/Fmax;p'] = 100
            
        for i in range(len(df5)):
            df5.loc[i, 'Fs'] = (bearing_capacity_ULS_shaft * df5.loc[i, 'Fs/Fmax;s']) / 100
            df5.loc[i, 'Fp'] = (bearing_capacity_ULS_tip * df5.loc[i, 'Fp/Fmax;p']) / 100
            df5.loc[i, 'Ftot'] = df5.loc[i, 'Fs'] + df5.loc[i, 'Fp']
            df5.loc[i, 'sel'] = 1000 * (l * df5.loc[i, 'Ftot'] + 0.5 * float(delta_l) * (df5.loc[i, 'Ftot'] + df5.loc[i, 'Fp'])) / (Epile * Apile)
            df5.loc[i, 'stotal-s1'] = df5.loc[i, 'sb'] + df5.loc[i, 'sel'] 
            df5.loc[i, 'Kv with sel'] = df5.loc[i, 'Ftot'] / df5.loc[i, 'stotal-s1']
            df5.loc[i, 'Kv without sel'] = df5.loc[i, 'Ftot'] / df5.loc[i, 'sb'] 
            
            df5 = df5.reindex(columns=['sb','sb/deq','Fs','Fp','Ftot','Fs/Fmax;s','Fp/Fmax;p','sel','stotal-s1','Kv with sel','Kv without sel'])
        
        interpolasi_with = interp1d(df5['Ftot'], df5['Kv with sel'])
        interpolasi_without = interp1d(df5['Ftot'], df5['Kv without sel'])
        
        kv_with = interpolasi_with(load_compression)
        kv_without = interpolasi_without(load_compression)
        
        kamus = {}
        kamus['kv with elastic settlement (MN/m)'] = [kv_with]
        kamus['kv without elastic settlement (MN/m)'] = [kv_without]
        
        df6 = pd.DataFrame(kamus)

        return df5, df6
    
class PileSummary():
    
    def __init__(self):
        self.name = "summary"
    
    def get_input_summary(self,
                          SF_compression,
                          SF_tension,
                          cohesiveapproach,
                          granularapproachshaft,
                          granularapproachendB,
                          pile_depth,
                          load_compression,
                          load_tension):
        
        label = ['SF compression',
                 'SF tension',
                 'Cohesive approach - shaft',
                 'Granular approach - shaft',
                 'Granular approach - end bearing',
                 'Pile depth (m)',
                 'Compression load (kN)',
                 'Tension_load (kN)']
        value = [SF_compression,
                 SF_tension,
                 cohesiveapproach,
                 granularapproachshaft,
                 granularapproachendB,
                 pile_depth,
                 load_compression,
                 load_tension]
        equal = ['=']*8
        kamus = {
            'Parameter' : label,
            '' : equal,
            'Value' : value
            }
        
        df = pd.DataFrame(kamus)
        
        return df
    
    def get_summary_multiple_data(self, result_capacity_table, result_kv_table, key_name, load_compression, load_tension):
        
        ndata = len(result_capacity_table)
        
        compression_capacity, tension_capacity, check_comp, check_tens, kvwi, kvwo = [],[],[],[],[],[]
        for i in range(ndata):
            summary_single = result_capacity_table[i]
            compression_capacity.append("%.2f"%(summary_single['Total'].iloc[1]))
            tension_capacity.append("%.2f"%(summary_single['Total'].iloc[3]))
            
            if load_compression < summary_single['Total'].iloc[1]:
                check_comp.append('OK')
            else:
                check_comp.append('NOT OK')
            
            if load_tension < summary_single['Total'].iloc[3]:
                check_tens.append('OK')
            else:
                check_tens.append('NOT OK')
            
            
            kv_table = result_kv_table[i]
            kvwi.append("%.2f"%(kv_table['kv with elastic settlement (MN/m)'].values[0]))
            kvwo.append("%.2f"%(kv_table['kv without elastic settlement (MN/m)'].values[0]))
            
        kamus1 = {
            'ID' : key_name,
            'Qall;comp_kN' : compression_capacity,
            'Qall;tens_kN' : tension_capacity,
            'check compression' : check_comp,
            'check tension' : check_tens,
            }
        
        kamus2 = {
            'ID' : key_name,
            'kv_with_sel_MN/m' : kvwi,
            'kv_w/o_sel_MN/m' : kvwo,
            }
        
        df1 = pd.DataFrame(kamus1)
        df2 = pd.DataFrame(kamus2)
        
        return df1, df2
            
class PileChart():
    
    def __init__(self):
        self.name = "chart_visualization"
    
    def get_bearing_capacity_chart(self, data_table, borehole_name):
        """
        Generate an interactive bearing capacity vs depth chart.

        Parameters
        ----------
        data_table : dataframe
            The dataframe generated using class PileCalculator and method calculate_bearing_capacity.
        borehole_name : string
            Name of borehole in particular.

        Returns
        -------
        fig : plotly chart
            
        """
                
        tabel = data_table

        x_var1 = tabel['compression_qall']
        x_var2 = tabel['qall_tens']
        y_var = tabel['bot']

        fig = go.Figure()
        fig.add_trace(go.Scatter(x = x_var1, y = y_var, mode = 'lines', name = 'Q_all;comp', marker_color='blue'))
        fig.add_trace(go.Scatter(x = x_var2, y = y_var, mode = 'lines', name = 'Q_all;tens', marker_color='red'))
        fig['layout']['yaxis']['autorange'] = "reversed"
        
        fig.update_layout(
            title='Bearing Capacity Chart (Qall) ' + borehole_name,
            xaxis_title='Pile bearing capacity (kN)',
            yaxis_title='Depth (m)',
            width=500,
            height=500,
            plot_bgcolor="white",
            font_family="Segoe UI",
            title_font_family = "Segoe UI semibold",
            margin=dict(l=30, r=20, t=30, b=30),
            )
        
        # fig.write_image("bearing_chart.png", scale=6, width=500, height=500)

        return fig
    
    def get_bearing_capacity_chart_mpl(self, data_table, borehole_name):
        """
        Generate a static bearing capacity vs depth chart.

        Parameters
        ----------
        data_table : dataframe
            The dataframe generated using class PileCalculator and method calculate_bearing_capacity.
        borehole_name : string
            Name of borehole in particular.

        Returns
        -------
        fig : matplotlib chart
            
        """
        
        tabel = data_table
        
        fig, ax = plt.subplots()

        x_var1 = tabel['compression_qall']
        x_var2 = tabel['qall_tens']
        y_var = tabel['bot']
        
        ax.plot(x_var1, y_var, color='blue', linestyle = 'solid', label=r'$Q_{all}$;comp')
        ax.plot(x_var2, y_var, color='green', linestyle = 'dashed', label=r'$Q_{all}$;tens')
        
        ax.set_xlim([0, None])
        ax.set_ylim([0, None])
        
        ax.invert_yaxis()
        ax.set_xlabel('Pile bearing capacity [kN]')
        ax.set_ylabel('Depth [m]')
        ax.legend()
        ax.grid(alpha=0.5)
        
        ax.set_title('Bearing Capacity ' + borehole_name)
        
        plt.savefig('chart_capacity_mpl.png', dpi=500)
        
        return fig
    
    def get_multi_bearing_capacity_chart(self, result_table, key_names, field):
        """
        Generate an interactive chart of bearing capacity vs depth for multiple outputs.

        Parameters
        ----------
        result_table : list
            List of dataframe, each dataframe is generated from class PileCalculator and method calculate_bearing_capacity.
        key_names : list
            List of labels (in string) associated with each of the result dataframe within the result_table.
        field : string
            Options = compression_ultimate, compression_allowed, tension_ultimate, tension_allowed.

        Returns
        -------
        fig : plotly chart
            
        """
        
        nchart = len(result_table)
        
        field_list = ["compression_ultimate", "compression_allowed", "tension_ultimate", "tension_allowed"]
        field_table = ["compression_qult", "compression_qall", "qult_tens", "qall_tens"]
        index_field = field_list.index(field)
        selected_field = field_table[index_field]
        
        fig = go.Figure()
        fig.update_layout(
            title='Bearing Capacity Chart (%s)'%(field),
            xaxis_title='Pile bearing capacity (kN)',
            yaxis_title='Depth (m)',
            width=500,
            height=500,
            plot_bgcolor="white",
            font_family="Segoe UI",
            title_font_family = "Segoe UI semibold",
            margin=dict(l=20, r=20, t=30, b=20),
            )
        fig['layout']['yaxis']['autorange'] = "reversed"
        
        for i in range(nchart):
            X = result_table[i][selected_field]
            Y = result_table[i]["bot"]
            
            fig.add_trace(go.Scatter(x = X, y = Y, mode = 'lines', name = key_names[i]))
        
        fig.write_image("chart_capacity_multi.png", scale=4, width=300, height=300)
        
        return fig
    
    def get_multi_bearing_capacity_chart_mpl(self, result_table, key_names, field, filename='chart_capacity_multi_mpl.png'):
        
        nchart = len(result_table)
        
        field_list = ["compression_ultimate", "compression_allowed", "tension_ultimate", "tension_allowed"]
        field_table = ["compression_qult", "compression_qall", "qult_tens", "qall_tens"]
        index_field = field_list.index(field)
        selected_field = field_table[index_field]
        
        fig, ax = plt.subplots()
        
        for i in range(nchart):
            
            tabel = result_table[i]

            x_var = tabel[selected_field]
            y_var = tabel['bot']
        
            ax.plot(x_var, y_var, linestyle = 'solid', label = key_names[i])
        
        ax.set_xlim([0, None])
        ax.set_ylim([0, None])
        
        ax.invert_yaxis()
        ax.set_xlabel('Pile bearing capacity [kN]')
        ax.set_ylabel('Depth [m]')
        ax.legend()
        ax.grid(alpha=0.5)
        
        ax.set_title('Bearing Capacity (%s)'%(field))
        
        plt.savefig(filename, dpi=500)
        
        return fig

class PDF(FPDF):
    def header(self):
        self.set_fill_color(30,101,84)
        self.rect(0,0,250,25,style='F')
        self.set_fill_color(38,119,103)
        self.rect(0,25,250,5,style='F')
        # Logo
        self.image(os.path.join(CODE_DIR,'wblogo_trans2.png'), 10, 8, 33)
        # Arial bold 15
        self.set_font('Helvetica', 'B', 15)
        self.set_text_color(250,250,250)
        # Move to the right
        self.cell(140)
        # Title
        self.cell(30, 10, 'Pile Bearing Capacity Result', 0, 0, 'C')
        # Line break
        self.ln(20)

    # Page footer
    def footer(self):
        # Position at 1.5 cm from bottom
        self.set_y(-15)
        # Arial italic 8
        self.set_font('Arial', 'I', 8)
        self.set_text_color(28,98,87)
        # Page number
        self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')
        
    def plot(self, plot, width):
        self.image(plot, x=15, y=45, w=width)
    
    def section_header(self, subheader):
        self.set_font('Arial', 'B', 20)
        self.set_text_color(30,101,84)
        self.multi_cell(w=180, h=20, txt=subheader, new_x='LEFT', new_y='NEXT', align='L', border=0)
    
    def table_dataframe(self, dataframe, colwidth, heading=True):
        df = dataframe
        df = df.applymap(str)  # Convert all data inside dataframe into string type
        columns = [list(df)]  # Get list of dataframe columns
        rows = df.values.tolist()  # Get list of dataframe rows
        data = columns + rows  # Combine columns and rows in one list
        self.set_font('Arial', '', 8)
        self.set_text_color(0,0,0)
        with self.table(borders_layout="SINGLE_TOP_LINE",
                        cell_fill_color=(230, 248, 246),  # grey
                        cell_fill_mode="ROWS",
                        line_height=self.font_size * 2.5,
                        text_align="CENTER",
                        col_widths=colwidth,
                        align='LEFT',
                        first_row_as_headings=heading,
                        width=190) as table:
            for data_row in data:
                row = table.row()
                for datum in data_row:
                    row.cell(datum)

class PileReport():
    def __init__(self):
        self.name = "report"
    
    def generate_report(self,
                        filename,
                        soil_table,
                        pile_catalog_table,
                        input_parameter_table,
                        selected_pile,
                        summary_capacity_table,
                        summary_kv_table,
                        kh_tables,
                        liquefaction_checkbox = False,
                        summary_capacity_liq = None
                        ):
        
        pdf = PDF('P', 'mm', 'A4')
        pdf.alias_nb_pages()
        
        pdf.add_page()
        pdf.section_header('Soil Profile')
        pdf.plot('chart_borelog_mpl.png', width=120)
        pdf.cell(w=100, h=120, new_x='LEFT', new_y='NEXT', align='L', border=0)
        pdf.section_header('Soil Table')
        pdf.table_dataframe(soil_table, colwidth=None)
        
        pdf.add_page()
        pdf.section_header('Pile Catalog')
        pdf.table_dataframe(pile_catalog_table, colwidth=None)
        
        pdf.add_page()
        pdf.section_header('Input Parameter')
        pdf.table_dataframe(input_parameter_table, colwidth=None)
        pdf.section_header('Selected Pile')
        pdf.table_dataframe(selected_pile, colwidth=None)
        
        pdf.add_page()
        pdf.section_header('Capacity Chart - compression (non-liquefied)')
        pdf.plot('chart_capacity_multi_mpl_CA.png', width=200)
        
        pdf.add_page()
        pdf.section_header('Capacity Chart - tension (non-liquefied)')
        pdf.plot('chart_capacity_multi_mpl_TA.png', width=200)
        
        pdf.add_page()
        pdf.section_header('Output Summary (non-liquefied)')
        pdf.table_dataframe(summary_capacity_table, colwidth=None)
        pdf.ln()
        pdf.table_dataframe(summary_kv_table, colwidth=None)
        
        pile_selected = selected_pile['Name'].tolist()
        for i in range(len(kh_tables)):
            pile_ID = pile_selected[i]
            kh_table = kh_tables[pile_ID]
            
            kh_table1 = kh_table[['Layer','kh lo [kN/m3]','kh med [kN/m3]','kh hi [kN/m3]']]
            kh_table2 = kh_table[['Layer','kh pile low/v2 [kN/m2]','kh pile low [kN/m2]','kh pile avg [kN/m2]','kh pile high [kN/m2]','kh pile * highv2 [kN/m2]']]
            
            pdf.add_page()
            pdf.section_header('Horizontal Spring Coefficient (non-liquefied) (%s)'%pile_ID)
            pdf.table_dataframe(kh_table1, colwidth=None)
            pdf.ln()
            pdf.table_dataframe(kh_table2, colwidth=None)
        
        if liquefaction_checkbox == True:
            pdf.add_page()
            pdf.section_header('Capacity Chart - compression (liquefied)')
            pdf.plot('chart_capacity_multi_mpl_liq_CA.png', width=200)
            
            pdf.add_page()
            pdf.section_header('Capacity Chart - tension (liquefied)')
            pdf.plot('chart_capacity_multi_mpl_liq_TA.png', width=200)
            
            pdf.add_page()
            pdf.section_header('Output Summary (liquefied)')
            pdf.table_dataframe(summary_capacity_liq, colwidth=None)
        
        pdf.output(filename, 'F')
        



        